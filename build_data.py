"""
build_data.py
=============

Reads the Sandbaggers League Excel workbook and writes a single public
"data.json" file that the website uses to show league results.

IMPORTANT PRIVACY RULE:
  This script NEVER copies player email addresses or phone numbers into
  data.json. Only golf results and public info go into the website.

How to run (from inside the repo folder):
    python3 build_data.py

By default it looks for a file whose name starts with
"Sandbaggers League Workbook" and ends in ".xlsx" in the current folder.
You can also pass the workbook path yourself:
    python3 build_data.py "My Workbook.xlsx"

We use openpyxl with data_only=True so we read the calculated VALUES
that Excel saved (for example 15.5), not the underlying formulas
(for example =E6+F6+G6).
"""

import glob
import json
import sys
from datetime import datetime

import openpyxl


# ---------------------------------------------------------------------------
# Small helper functions
# ---------------------------------------------------------------------------

def find_workbook():
    """Return the path to the workbook file.

    If the user passed a filename on the command line, use that.
    Otherwise look for any "Sandbaggers League Workbook*.xlsx" in this folder.
    """
    if len(sys.argv) > 1:
        return sys.argv[1]

    matches = sorted(glob.glob("Sandbaggers League Workbook*.xlsx"))
    if not matches:
        raise FileNotFoundError(
            "Could not find a 'Sandbaggers League Workbook*.xlsx' file. "
            "Pass the workbook path as an argument, e.g.:\n"
            '    python3 build_data.py "Sandbaggers League Workbook.xlsx"'
        )
    return matches[0]


def money(value):
    """Turn a cell into a number, treating blanks as 0.

    Blank money cells (None or empty text) become 0. Everything else is
    converted to a float, and whole numbers like 8.0 become 8.
    """
    if value is None or value == "":
        return 0
    number = float(value)
    # Show 8 instead of 8.0, but keep 15.5 as 15.5.
    return int(number) if number == int(number) else number


def text(value):
    """Return a clean string, or "" if the cell is blank."""
    if value is None:
        return ""
    return str(value).strip()


def as_date(value):
    """Format a date cell as 'YYYY-MM-DD', or "" if there is no date."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if value is None or value == "":
        return ""
    # If it is already text, just hand it back trimmed.
    return str(value).strip()


def row_is_empty(row):
    """True if every cell in the row is blank."""
    return all(cell is None or cell == "" for cell in row)


# ---------------------------------------------------------------------------
# One function per sheet. Each returns plain Python data.
# ---------------------------------------------------------------------------

def build_league():
    """Fixed league info. This never changes, so we just hard-code it."""
    return {
        "name": "Sandbaggers",
        "course": "Viera East Golf Club",
        "tees": "Green",
        "holes": 9,
        "night": "Thursday",
    }


def build_players(wb):
    """Read the Roster sheet.

    Header row is row 3:
      # | Player Name | Email | Phone | GHIN | ...handicaps... | Status | Notes
    We keep ONLY name, ghin, and status. Email and phone are ignored on
    purpose so they never reach the website.
    """
    ws = wb["Roster"]
    players = []
    # Data starts on row 4 (rows 1-2 are title/instructions, row 3 is headers).
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row_is_empty(row):
            continue
        name = text(row[1])   # column B = Player Name
        if not name:
            continue          # skip any stray rows with no name
        players.append({
            "name": name,
            "ghin": row[4],   # column E = GHIN index (a number like 12.8)
            "status": text(row[9]),  # column J = Status (Active / etc.)
        })
    return players


def build_standings(wb):
    """Read the Season Standings sheet and sort by total winnings (high to low).

    Header row is row 5:
      Rank | Player | Weeks Played | Team Game $ | Skins $ | CTP $ | Total Winnings $
    We skip blank rows and the "LEAGUE TOTAL" summary row at the bottom.
    """
    ws = wb["Season Standings"]
    standings = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row_is_empty(row):
            continue
        player = text(row[1])   # column B = Player
        # Skip the summary row and any nameless rows.
        if not player or player.upper() == "LEAGUE TOTAL":
            continue
        standings.append({
            "rank": row[0],                 # column A = Rank (may be blank)
            "player": player,
            "weeksPlayed": money(row[2]),   # column C
            "teamGame": money(row[3]),      # column D = Team Game $
            "skins": money(row[4]),         # column E = Skins $
            "ctp": money(row[5]),           # column F = CTP $
            "total": money(row[6]),         # column G = Total Winnings $
        })

    # Sort by total winnings, highest first.
    standings.sort(key=lambda p: p["total"], reverse=True)
    return standings


def build_weekly_results(wb):
    """Read the Results Log sheet (one row per player per week).

    Header row is row 5:
      Wk # | Date | Format | Player | Team game $ | Skins $ | CTP $ | Total $ | Notes
    Rows without a player name are placeholder/empty rows and are skipped.
    """
    ws = wb["Results Log"]
    results = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row_is_empty(row):
            continue
        player = text(row[3])   # column D = Player
        if not player:
            continue            # skip placeholder rows that only have week/date
        results.append({
            "week": row[0],               # column A = Wk #
            "date": as_date(row[1]),      # column B = Date
            "format": text(row[2]),       # column C = Format
            "player": player,
            "teamGame": money(row[4]),    # column E = Team game $
            "skins": money(row[5]),       # column F = Skins $
            "ctp": money(row[6]),         # column G = CTP $
            "total": money(row[7]),       # column H = Total $
        })
    return results


def build_schedule(wb):
    """Read the Schedule sheet (one row per week of the season).

    Header row is row 3:
      Wk # | Date | Format | Tees | CTP Hole #1 | CTP Hole #2 | Notes | Status
    """
    ws = wb["Schedule"]
    schedule = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row_is_empty(row):
            continue
        if row[0] is None:      # must have a week number to be a real week
            continue
        schedule.append({
            "week": row[0],             # column A = Wk #
            "date": as_date(row[1]),    # column B = Date
            "format": text(row[2]),     # column C = Format
            "tees": text(row[3]),       # column D = Tees
            "ctp1": row[4],             # column E = CTP Hole #1 (hole number)
            "ctp2": row[5],             # column F = CTP Hole #2 (hole number)
            "status": text(row[7]),     # column H = Status
        })
    return schedule


def build_course(wb):
    """Read the Course Info sheet into a front nine and back nine.

    Layout (rows):
      Front nine -> row 6 = Hole numbers, row 7 = Par, row 8 = Stroke Index
      Back nine  -> row 12 = Hole numbers, row 13 = Par, row 14 = Stroke Index
    Columns B..J (indexes 1..9) hold the nine holes; column K is the Total.
    """
    ws = wb["Course Info"]
    rows = list(ws.iter_rows(values_only=True))

    def nine(hole_row, par_row, index_row):
        """Build a list of 9 holes from three rows of the sheet."""
        holes = []
        # Columns index 1..9 are the nine holes (skip column A label and Total).
        for col in range(1, 10):
            holes.append({
                "hole": hole_row[col],
                "par": par_row[col],
                "strokeIndex": index_row[col],
            })
        return holes

    # rows is 0-based, so sheet row 6 is rows[5], row 7 is rows[6], etc.
    front = nine(rows[5], rows[6], rows[7])
    back = nine(rows[11], rows[12], rows[13])
    return {"frontNine": front, "backNine": back}


def build_payouts(wb):
    """Read the Payout Table sheet.

    Header row is row 3:
      Players | Pot ($3 ea) | 60% (1st calc) | 40% (2nd calc) | 1st Place $ | 2nd Place $
    We keep the number of players and the actual 1st / 2nd place dollar amounts.
    """
    ws = wb["Payout Table"]
    payouts = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row_is_empty(row):
            continue
        # A real payout row starts with a player count (a number).
        if not isinstance(row[0], (int, float)):
            continue
        payouts.append({
            "players": row[0],          # column A = number of players
            "first": money(row[4]),     # column E = 1st Place $
            "second": money(row[5]),    # column F = 2nd Place $
        })
    return payouts


# ---------------------------------------------------------------------------
# Main: read everything, write data.json
# ---------------------------------------------------------------------------

def main():
    workbook_path = find_workbook()
    print(f"Reading workbook: {workbook_path}")

    # data_only=True -> read saved calculated values, not the formulas.
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    data = {
        "league": build_league(),
        "players": build_players(wb),
        "standings": build_standings(wb),
        "weeklyResults": build_weekly_results(wb),
        "schedule": build_schedule(wb),
        "course": build_course(wb),
        "payouts": build_payouts(wb),
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    # Plain-English summary so you can confirm the mapping looks right.
    weeks_played = sorted({r["week"] for r in data["weeklyResults"]})
    print("\nWrote data.json. Quick summary:")
    print(f"  Players on roster:      {len(data['players'])}")
    print(f"  Players in standings:   {len(data['standings'])}")
    print(f"  Weeks in schedule:      {len(data['schedule'])}")
    print(f"  Weeks with results:     {len(weeks_played)}  ({weeks_played})")
    print(f"  Result rows (player-weeks): {len(data['weeklyResults'])}")
    print(f"  Payout rows:            {len(data['payouts'])}")


if __name__ == "__main__":
    main()
